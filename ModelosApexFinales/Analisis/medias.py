import pandas as pd
from openpyxl import load_workbook

archivo_excel = "resultados_tensores.xlsx"

# Cargar el archivo existente
wb = load_workbook(archivo_excel)
resumen = []

for nombre_hoja in wb.sheetnames:
    if nombre_hoja == "ResumenMedias":
        continue  # Evitar duplicar si ya existe

    ws = wb[nombre_hoja]
    if ws.max_row < 2:
        continue  # Saltar hojas vacías o incompletas

    fila_media = [cell.value for cell in ws[ws.max_row]]
    resumen.append([nombre_hoja] + fila_media[1:])  # Quitar "Media"

# Definir columnas
columnas_base = [
    "El. Pos. ID","von Mises",
    "X Component", "Y Component", "Z Component",
    "XY Component", "YZ Component", "ZX Component",
    "XY Engr. Component", "YZ Engr. Component", "ZX Engr. Component",
    "SigmaVonMises2D"
]
columnas_resumen = ["Archivo"] + columnas_base

# Crear DataFrame resumen
df_resumen = pd.DataFrame(resumen, columns=columnas_resumen)

# Escribir la hoja resumen en el Excel existente
with pd.ExcelWriter(archivo_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df_resumen.to_excel(writer, sheet_name="ResumenMedias", index=False)

print("✅ Hoja 'ResumenMedias' añadida correctamente.")
