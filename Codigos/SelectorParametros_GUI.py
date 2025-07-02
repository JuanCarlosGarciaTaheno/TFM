import tkinter as tk
import os
from tkinter import ttk, messagebox
import pandas as pd
import openpyxl  
from PyQt5.QtWidgets import QMessageBox

# Diccionario de pasos por métrica
# Diccionario de pasos estándar por métrica
pasos_por_metrica = {
    1: 0.25, 1.1: 0.25, 1.2: 0.25, 1.4: 0.3, 1.6: 0.35, 1.8: 0.35, 2: 0.4, 2.2: 0.45,
    2.5: 0.45, 3: 0.5, 3.5: 0.6, 4: 0.7, 4.5: 0.75, 5: 0.8, 6: 1.0, 7: 1.0, 8: 1.25,
    9: 1.25, 10: 1.5, 11: 1.5, 12: 1.75, 14: 2.0, 16: 2.0, 18: 2.5, 20: 2.5
}



# === Configuración Inicial ===
ventana = tk.Tk()
ventana.title("Selector de Parámetros")
ventana.geometry("900x800")

# Rutas de archivos
ruta_excel = "C:/Users/jc.garcia-taheno/Desktop/CE_3_TahenoHijes/Param_Union_TornHelic.xlsx"


# Cargar datos
df = pd.read_excel(ruta_excel, sheet_name="Long_Roscado")
df_helicoil = pd.read_excel(ruta_excel, sheet_name="Helicoil_Catia")  

# Obtener valores únicos de D1Hc_max de la tabla de Helicoils
valores_d1hc_max = sorted(df_helicoil['D1Hc_max'].unique())

valores_W = sorted(df_helicoil['W_mm'].unique())


# === Contenedores ===
contenedor = tk.Frame(ventana, bg='white', bd=2)
contenedor.place(relx=0.5, rely=0.5, anchor="center")

frame_numericos = tk.LabelFrame(contenedor, text="Parámetros Numéricos", bg="white")
frame_numericos.grid(row=0, column=0, padx=10, pady=10)

frame_opciones = tk.LabelFrame(contenedor, text="Parámetros de Opción", bg="white")
frame_opciones.grid(row=0, column=1, padx=10, pady=10)

# === Diccionarios y Contadores ===
parametros = {}
fila_numericos = 0
fila_opciones = 0

# === Función para agregar campos ===
def agregar_parametro(frame, nombre, tipo, opciones=None, fila=None):
    tk.Label(frame, text=nombre, bg="white").grid(row=fila, column=0, sticky="e", pady=2)
    if tipo == "entry":
        entrada = tk.Entry(frame)
        entrada.grid(row=fila, column=1, pady=2)
        parametros[nombre] = entrada
    elif tipo == "combo" and opciones is not None:
        combo = ttk.Combobox(frame, values=opciones, state="readonly")
        combo.grid(row=fila, column=1, pady=2)
        # Desactivar interacción solo para el campo "W"
        if nombre == "W":
            combo.configure(state="disabled")
        parametros[nombre] = combo

# === Parámetros Numéricos ===
parametros_numericos = [
    ("Metrica_Seleccionada", "combo", ['M2','M2.5', 'M3','M3.5', 'M4', 'M5', 'M6','M7', 'M8','M9', 'M10', 'M11','M12','M14','M16']),
    ("Paso_Seleccionado", "entry"),
    ("t2_Seleccionado", "combo", ["1.0 d", "1.5 d", "2.0 d", "2.5 d", "3.0 d"]),
    ("D1HC_max_seleccionado", "combo", valores_d1hc_max), 
    ("W", "combo", valores_W),   
    ("Espesor_PiezaSuperior", "entry"),
    ("Espesor_PiezaInferior", "entry"),
    ("longitud_Seleccionada", "combo", [str(x) for x in [2.5, 3, 4, 5, 6, 8, 10, 12, 16, 20, 25, 30, 35, 40,
        45, 50, 55, 60, 65, 70, 80, 90, 100, 110, 120, 130, 140, 150, 160, 180, 200, 220, 240, 260, 280, 300]])
]

for nombre, tipo, *opciones in parametros_numericos:
    agregar_parametro(frame_numericos, nombre, tipo, opciones[0] if opciones else None, fila_numericos)
    fila_numericos += 1



# Función para actualizar D1HC_max según métrica seleccionada
def actualizar_d1hc_max(*args):
    metrica = parametros["Metrica_Seleccionada"].get()
    if metrica:
        # Extraer el número de la métrica (ej: "M3" -> 3)
        try:
            num_metrica = float(metrica.replace("M", "").replace("_", "."))
        except ValueError:
            return
        
        # Filtrar valores de D1Hc_max para esta métrica
        valores_filtrados = df_helicoil[df_helicoil['Metrica'] == num_metrica]['D1Hc_max'].unique()
        valores_filtrados = sorted(valores_filtrados)
        
        # Actualizar combobox
        parametros["D1HC_max_seleccionado"]['values'] = valores_filtrados
        if valores_filtrados:
            parametros["D1HC_max_seleccionado"].set(valores_filtrados[0])  # Establecer primer valor por defecto

# Vincular evento de cambio en la métrica
parametros["Metrica_Seleccionada"].bind("<<ComboboxSelected>>", actualizar_d1hc_max)


# Función para actualizar W según métrica seleccionada
def actualizar_W(*args):
    metrica = parametros["Metrica_Seleccionada"].get()
    if metrica:
        # Extraer el número de la métrica (ej: "M3" -> 3)
        try:
            num_metrica = float(metrica.replace("M", "").replace("_", "."))
        except ValueError:
            return
        
        # Filtrar valores de W para esta métrica
        valores_filtrados = df_helicoil[df_helicoil['Metrica'] == num_metrica]['W_mm'].unique()
        valores_filtrados = sorted(valores_filtrados)
        
        # Actualizar combobox
        parametros["W"]['values'] = valores_filtrados
        if valores_filtrados:
            parametros["W"].set(valores_filtrados[0])  # Establecer primer valor por defecto

# Vincular evento de cambio en la métrica
parametros["Metrica_Seleccionada"].bind("<<ComboboxSelected>>", actualizar_W, actualizar_d1hc_max)

# Función para sincronizar la selección entre t2_Seleccionado y W, considerando la métrica seleccionada
def sincronizar_seleccion_t2_w(*args):
    t2_seleccionado = parametros["t2_Seleccionado"].get()  # Obtener el valor seleccionado de T2
    metrica_seleccionada = parametros["Metrica_Seleccionada"].get()  # Obtener la métrica seleccionada
    
    if not metrica_seleccionada:
        return  # Si no hay métrica seleccionada, no hacer nada
    
    # Extraer el número de la métrica seleccionada
    try:
        num_metrica = float(metrica_seleccionada.replace("M", "").replace("_", "."))
    except ValueError:
        return
    
    # Filtrar los valores de W para esta métrica
    valores_filtrados_W = df_helicoil[df_helicoil['Metrica'] == num_metrica]['W_mm'].unique()
    valores_filtrados_W = sorted(valores_filtrados_W)
    
    # Lista de opciones disponibles para T2
    opciones_t2 = ["1.0 d", "1.5 d", "2.0 d", "2.5 d", "3.0 d"]
    
    # Verificar si el valor de T2 está dentro de las opciones disponibles
    if t2_seleccionado in opciones_t2:
        indice_t2 = opciones_t2.index(t2_seleccionado)
        
        # Asegurarse de que el índice no se salga del rango de W para la métrica seleccionada
        if indice_t2 < len(valores_filtrados_W):
            parametros["W"].set(valores_filtrados_W[indice_t2])  # Establecer el valor de W correspondiente para esta métrica

# Vincular evento de cambio en t2_Seleccionado
parametros["t2_Seleccionado"].bind("<<ComboboxSelected>>", sincronizar_seleccion_t2_w)

# Función para actualizar W según métrica seleccionada
def actualizar_W(*args):
    metrica = parametros["Metrica_Seleccionada"].get()
    if metrica:
        # Extraer el número de la métrica (ej: "M3" -> 3)
        try:
            num_metrica = float(metrica.replace("M", "").replace("_", "."))
        except ValueError:
            return
        
        # Filtrar valores de W para esta métrica
        valores_filtrados = df_helicoil[df_helicoil['Metrica'] == num_metrica]['W_mm'].unique()
        valores_filtrados = sorted(valores_filtrados)
        
        # Actualizar combobox
        parametros["W"]['values'] = valores_filtrados
        if valores_filtrados:
            parametros["W"].set(valores_filtrados[0])  # Establecer primer valor por defecto
            # Función para actualizar el paso según la métrica seleccionada

    metrica = parametros["Metrica_Seleccionada"].get()
    if metrica:
        try:
            num_metrica = float(metrica.replace("M", "").replace("_", "."))
        except ValueError:
            return
        
        paso = pasos_por_metrica.get(num_metrica)
        if paso is not None:
            parametros["Paso_Seleccionado"].delete(0, tk.END)
            parametros["Paso_Seleccionado"].insert(0, str(paso))

# Vincular evento de cambio en la métrica
parametros["Metrica_Seleccionada"].bind("<<ComboboxSelected>>", actualizar_W, actualizar_d1hc_max)










# Campos de solo lectura
for nombre in ["ls", "lg"]:
    agregar_parametro(frame_numericos, nombre, "entry", fila=fila_numericos)
    parametros[nombre].config(state="readonly")
    fila_numericos += 1



# === Selector General de Modo ===
tk.Label(frame_opciones, text="Modo General de Selección:", bg="white", font=("Arial", 10, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", pady=(10, 2))
modo_general = ttk.Combobox(frame_opciones, values=["Máxima Separación", "Mínima Separación", "Media Separación"], state="readonly")
modo_general.grid(row=1, column=0, columnspan=2, pady=(0, 10))
modo_general.set("Seleccionar Modo")
ttk.Separator(frame_opciones, orient="horizontal").grid(row=2, column=0, columnspan=2, sticky="ew", pady=(0, 10))

fila_opciones = 3  # Comenzamos tras el separador

# === Diccionario de Opciones ===
opciones_dict = {
    "Opcion_d": ["d_max", "d_min", "d_media"],
    "Opcion_d3": ["d3_max", "d3_min", "d3_media"],
    "Opcion_D1": ["D1_max", "D1_min", "D1_media"],
    "Opcion_D2": ["D2_max", "D2_min", "D2_media"],
 #   "Opcion_d1": ["d1_max", "d1_min", "d1_media"],
    "Opcion_D1HC": ["D1HC_max", "D1HC_min", "D1HC_media"],
    "Opcion_d2": ["d2_max", "d2_min", "d2_media"],
    "Opcion_d1InteriorArandela": ["d1_interior_max", "d1_interior_min","d1_int_medio"],
    "Opcion_d2ExteriorArandela": ["d2_exterior_nominal", "d2_ext_min","d2_ext_medio"],
    "Opcion_dk": ["dk_max", "dk_min", "dk_max_grooved_head","dk_media"],
    "Opcion_k": ["k_max", "k_min","k_media"],
    "Opcion_hArandela": ["h_max", "h_min", "h_nominal"],
    "Opcion_ProfRosca_e1": ["e1_Normal", "e1_Short", "e1_Long"],
    "Opcion_Longitud": ["l_nominal", "l_min", "l_max"]
}

# === Secciones por Categoría ===
secciones = {
    # "Parámetros de Rosca": [
    #     "Opcion_d", "Opcion_d3", "Opcion_D1", "Opcion_D2", "Opcion_d1", "Opcion_D1HC", "Opcion_d2"
    # ],
    "Parámetros de Rosca": [
        "Opcion_d", "Opcion_d3", "Opcion_D1", "Opcion_D2",  "Opcion_D1HC", "Opcion_d2"
    ],
    "Parámetros de Arandela": [
        "Opcion_d1InteriorArandela", "Opcion_d2ExteriorArandela", "Opcion_hArandela"
    ],
    "Parámetros de Cabeza": [
       "Opcion_dk", "Opcion_k"
    ],
    "Parámetros de Longitudes": [
        "Opcion_ProfRosca_e1", "Opcion_Longitud"
    ]
}

for titulo, claves in secciones.items():
    tk.Label(frame_opciones, text=titulo, bg="white", font=("Arial", 10, "bold")).grid(row=fila_opciones, column=0, columnspan=2, sticky="w", pady=(10, 2))
    fila_opciones += 1
    for clave in claves:
        agregar_parametro(frame_opciones, clave, "combo", opciones_dict[clave], fila_opciones)
        fila_opciones += 1


def actualizar_parametros(*args):
    seleccion = modo_general.get()
    if seleccion == "Mínima Separación":
        valores = {
            "Opcion_d": "d_max", "Opcion_d3": "d3_max", "Opcion_D1": "D1_min",
            "Opcion_D2": "D2_min",  "Opcion_D1HC": "D1HC_min", "Opcion_d2": "d2_max"
        }
    elif seleccion == "Máxima Separación":
        valores = {
            "Opcion_d": "d_min", "Opcion_d3": "d3_min", "Opcion_D1": "D1_max",
            "Opcion_D2": "D2_max",  "Opcion_D1HC": "D1HC_max", "Opcion_d2": "d2_min"
        }
    elif seleccion == "Media Separación":
        valores = {
            "Opcion_d": "d_media", "Opcion_d3": "d3_media", "Opcion_D1": "D1_media",
            "Opcion_D2": "D2_media",  "Opcion_D1HC": "D1HC_media", "Opcion_d2": "d2_media"
        }
    else:
        return
    
    for k, v in valores.items():
        if k in parametros:
            parametros[k].set(v)
modo_general.bind("<<ComboboxSelected>>", actualizar_parametros)

# Función para actualizar ls y lg
def actualizar_ls_lg():
    metrica_seleccionada = parametros["Metrica_Seleccionada"].get()
    longitud_valor = parametros["longitud_Seleccionada"].get()

    if not metrica_seleccionada or not longitud_valor:
        print("Debe seleccionar una métrica y una longitud.")
        return

    try:
        longitud_seleccionada = float(longitud_valor)
    except ValueError:
        print("La longitud seleccionada no es válida.")
        return

    longitud_datos = df[df['l_nominal'] == longitud_seleccionada]
    numero_metrica = metrica_seleccionada.strip("M")
    metrica_ls = f"M{numero_metrica}_ls_min"
    metrica_lg = f"M{numero_metrica}_lg_max"

    if metrica_ls in df.columns:
        valor_ls = longitud_datos[metrica_ls].values[0]
        parametros["ls"].config(state="normal")
        parametros["ls"].delete(0, tk.END)
        parametros["ls"].insert(0, valor_ls)
        parametros["ls"].config(state="readonly")
    else:
        print(f"Columna {metrica_ls} no encontrada.")
        parametros["ls"].config(state="normal")
        parametros["ls"].delete(0, tk.END)
        parametros["ls"].config(state="readonly")

    if metrica_lg in df.columns:
        valor_lg = longitud_datos[metrica_lg].values[0]
        parametros["lg"].config(state="normal")
        parametros["lg"].delete(0, tk.END)
        parametros["lg"].insert(0, valor_lg)
        parametros["lg"].config(state="readonly")
    else:
        print(f"Columna {metrica_lg} no encontrada.")
        parametros["lg"].config(state="normal")
        parametros["lg"].delete(0, tk.END)
        parametros["lg"].config(state="readonly")

# Botón para actualizar ls y lg
btn_actualizar = tk.Button(contenedor, text="Actualizar ls y lg", command=actualizar_ls_lg, bg="#4CAF50", fg="white")
btn_actualizar.grid(row=1, column=0, columnspan=2, pady=10)




comentario_W = tk.Label(frame_numericos,
    text=" Todas las unidades serán expresadas en milímetros y con punto decimal.",
    fg="black", bg="white", wraplength=400, justify="left")
comentario_W.grid(row=fila_numericos, column=0, columnspan=2, pady=(10, 0), sticky="w")
fila_numericos += 1
# === Comentario en la GUI sobre la condición de W ===
comentario_W = tk.Label(frame_numericos,
    text="⚠️ W (longitud del Helicoil en mm) debe ser mayor que la longitud seleccionada [mm] y el espesor de la pieza inferior [mm]. Además, para métricas M14 y M16 no dispondrán de valores para t 3.0d, ver Tabla Helicoil.",
    fg="orange", bg="white", wraplength=400, justify="left")
comentario_W.grid(row=fila_numericos, column=0, columnspan=2, pady=(10, 0), sticky="w")
fila_numericos += 1




# Función para guardar en Excel
def guardar_en_excel():
    nombre_archivo = "C:/Users/jc.garcia-taheno/Desktop/CE_3_TahenoHijes/Codigos/parametros_seleccionados.xlsx"

    # Crear el archivo si no existe
    if os.path.exists(nombre_archivo):
        wb = openpyxl.load_workbook(nombre_archivo)
    else:
        wb = openpyxl.Workbook()

    # Hoja 1: Parámetros Numéricos (incluyendo ls y lg)
    if "1_ParamNum" not in wb.sheetnames:
        ws1 = wb.create_sheet("1_ParamNum", 0)
    else:
        ws1 = wb["1_ParamNum"]
        
    for col, (nombre, tipo, *opciones) in enumerate(parametros_numericos, 1):
        ws1.cell(row=1, column=col, value=nombre)  # Escribir nombre de parámetro en la primera fila

    for col, (nombre, tipo, *opciones) in enumerate(parametros_numericos, 1):
        valor = parametros[nombre].get() if isinstance(parametros[nombre], (tk.Entry, ttk.Combobox)) else ""
        ws1.cell(row=2, column=col, value=valor)  # Escribir valor en la segunda fila

    # Hoja 2: Parámetros de Rosca
    if "2_Rosca" not in wb.sheetnames:
        ws2 = wb.create_sheet("2_Rosca")
    else:
        ws2 = wb["2_Rosca"]

    for col, nombre in enumerate(secciones["Parámetros de Rosca"], 1):
        ws2.cell(row=1, column=col, value=nombre)  # Escribir nombre de parámetro en la primera fila

    for col, nombre in enumerate(secciones["Parámetros de Rosca"], 1):
        valor = parametros.get(nombre, "").get() if nombre in parametros else ""
        ws2.cell(row=2, column=col, value=valor)  # Escribir valor en la segunda fila

    # Hoja 3: Parámetros de Arandela
    if "3_Arandela" not in wb.sheetnames:
        ws3 = wb.create_sheet("3_Arandela")
    else:
        ws3 = wb["3_Arandela"]

    for col, nombre in enumerate(secciones["Parámetros de Arandela"], 1):
        ws3.cell(row=1, column=col, value=nombre)  # Escribir nombre de parámetro en la primera fila

    for col, nombre in enumerate(secciones["Parámetros de Arandela"], 1):
        valor = parametros.get(nombre, "").get() if nombre in parametros else ""
        ws3.cell(row=2, column=col, value=valor)  # Escribir valor en la segunda fila

    # Hoja 4: Parámetros de Cabeza
    if "4_Cabeza" not in wb.sheetnames:
        ws3 = wb.create_sheet("4_Cabeza")
    else:
        ws3 = wb["4_Cabeza"]

    for col, nombre in enumerate(secciones["Parámetros de Cabeza"], 1):
        ws3.cell(row=1, column=col, value=nombre)  # Escribir nombre de parámetro en la primera fila

    for col, nombre in enumerate(secciones["Parámetros de Cabeza"], 1):
        valor = parametros.get(nombre, "").get() if nombre in parametros else ""
        ws3.cell(row=2, column=col, value=valor)  # Escribir valor en la segunda fila

    # Hoja 5: Parámetros de Longitudes
    if "5_Longitud" not in wb.sheetnames:
        ws4 = wb.create_sheet("5_Longitud")
    else:
        ws4 = wb["5_Longitud"]

    for col, nombre in enumerate(secciones["Parámetros de Longitudes"], 1):
        ws4.cell(row=1, column=col, value=nombre)  # Escribir nombre de parámetro en la primera fila

    for col, nombre in enumerate(secciones["Parámetros de Longitudes"], 1):
        valor = parametros.get(nombre, "").get() if nombre in parametros else ""
        ws4.cell(row=2, column=col, value=valor)  # Escribir valor en la segunda fila


        # Hoja 6: Valores de ls y lg
    if "6_ls_lg" not in wb.sheetnames:
        ws6 = wb.create_sheet("6_ls_lg")
    else:
        ws6 = wb["6_ls_lg"]

    # Escribir encabezados
    ws6.cell(row=1, column=1, value="ls (mínimo)")
    ws6.cell(row=1, column=2, value="lg (máximo)")

    # Obtener los valores actuales desde la interfaz
    ls = parametros["ls"].get()
    lg = parametros["lg"].get()

    # Escribir los valores
    ws6.cell(row=2, column=1, value=ls)
    ws6.cell(row=2, column=2, value=lg)


    # Guardar el archivo
    wb.save(nombre_archivo)
    print("Parámetros guardados en el Excel.")

# --- Botón de guardar en Excel ---
btn_guardar = tk.Button(contenedor, text="Guardar en Excel", command=guardar_en_excel, bg="#4CAF50", fg="white")
btn_guardar.grid(row=2, column=0, columnspan=2, pady=10)

# Ejecutar la ventana
ventana.mainloop()